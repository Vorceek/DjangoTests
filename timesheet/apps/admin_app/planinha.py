import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.base_app.models import Cliente, Servico, Atividade

def exportar_hierarquia_para_excel(request):
    # Criando um novo arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hierarquia"

    # Adicionando o mês no canto direito da primeira linha (Coluna A)
    mes = "Fevereiro"
    ws["A1"] = mes
    ws["A1"].font = Font(name="Aptos Narrow", size=9, bold=True)
    ws["A1"].alignment = Alignment(horizontal="right")

    # Primeira linha da Coluna B - Campo em branco
    ws["B1"] = ""

    # Segunda linha da Coluna B - "TOTAL" em negrito
    ws["B2"] = "TOTAL"
    ws["B2"].font = Font(name="Aptos Narrow", size=9, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center")

    # Adiciona uma linha em branco na Coluna A
    ws.append(["", ""])

    # Cor de fundo para clientes (empresas)
    cliente_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fonte padrão para os elementos
    default_font = Font(name="Aptos Narrow", size=9)
    bold_font = Font(name="Aptos Narrow", size=9, bold=True)

    row = 4  # Começa na quarta linha (primeira é o mês, segunda é "TOTAL", terceira é espaço)

    # Percorre os dados e insere na planilha
    for cliente in Cliente.objects.all().order_by("nome"):  # Ordena os clientes alfabeticamente
        ws.append([cliente.nome, ""])  # Adiciona a empresa (cliente) na Coluna A, Coluna B vazia
        ws[f"A{row}"].font = bold_font
        ws[f"A{row}"].fill = cliente_fill  # Fundo cinza claro para empresas
        row += 1

        servicos = Servico.objects.filter(clientes=cliente).order_by("nome")  # Ordena os serviços
        for servico in servicos:
            ws.append([f"    {servico.nome}", ""])  # Recuo para serviços
            ws[f"A{row}"].font = bold_font
            row += 1

            atividades = Atividade.objects.filter(servicos=servico).order_by("nome")  # Ordena as atividades
            for atividade in atividades:
                ws.append([f"        {atividade.nome}", ""])  # Recuo maior para atividades
                ws[f"A{row}"].font = default_font  # Fonte padrão
                row += 1

    # Ajusta a largura das colunas para melhor visualização
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10

    # Salvar no buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Retornar resposta para download
    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="hierarquia_atividades.xlsx"'
    
    return response
