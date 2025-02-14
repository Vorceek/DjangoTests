from django.http import HttpResponse
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.base_app.permissions import AdminRequiredMixin, aside_icons
from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd
import pytz
from django.views import View
from django.shortcuts import render
from django.core.paginator import Paginator
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from apps.atividade_app.models import Cliente, RegistroAtividadeModel, Servico, Atividade
from django.contrib.auth.models import User
from django.utils.timezone import make_naive
from django.db.models import Sum, F
from django.utils.timezone import now
from django.db.models.functions import ExtractWeekDay
import json
from django.db.models import F, ExpressionWrapper, fields, Case, When, Value
from django.db.models.functions import Coalesce

class AdminView(LoginRequiredMixin, TemplateView):
    template_name = "admin/hub.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        fixed_groups = ["ADMINISTRADOR", "USER"]

        # Obtém os setores dinâmicos (excluindo os fixos)
        setores_dinamicos = list(
            self.request.user.groups.exclude(name__in=fixed_groups).values_list('name', flat=True)
        )

        if not setores_dinamicos:
            context.update({
                "atividades_por_dia": json.dumps([]),
                "servicos_labels": json.dumps([]),
                "servicos_values": json.dumps([]),
                "barras_labels": json.dumps([]),
                "barras_datasets": json.dumps([]),
            })
            return context

        # Obtém as datas do filtro
        data_inicio_str = self.request.GET.get("data_inicio")
        data_fim_str = self.request.GET.get("data_fim")

        if data_inicio_str and data_fim_str:
            try:
                data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
                data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
            except ValueError:
                data_inicio = now().date() - timedelta(days=now().weekday() + 7)  # Segunda-feira passada
                data_fim = data_inicio + timedelta(days=6)  # Domingo passado
        else:
            data_inicio = now().date() - timedelta(days=now().weekday() + 7)  # Segunda-feira passada
            data_fim = data_inicio + timedelta(days=6)  # Domingo passado

        # ---- Gráfico de atividades por dia ----
        atividades = RegistroAtividadeModel.objects.filter(
            RAM_colaborador__groups__name__in=setores_dinamicos,
            RAM_dataInicial__date__gte=data_inicio,
            RAM_dataInicial__date__lte=data_fim
        ).annotate(dia_semana=ExtractWeekDay("RAM_dataInicial")).values("dia_semana")

        atividades_count = {i: 0 for i in range(1, 8)}  # Domingo (1) até Sábado (7)
        for atividade in atividades:
            dia = atividade["dia_semana"]
            if dia in atividades_count:
                atividades_count[dia] += 1

        atividades_por_dia = [atividades_count.get(i, 0) for i in range(1, 8)]

        # ---- Gráfico de horas por serviço (rosquinha) ----
        servicos_atividades = (
            RegistroAtividadeModel.objects
            .filter(
                RAM_colaborador__groups__name__in=setores_dinamicos,
                RAM_dataInicial__date__gte=data_inicio,
                RAM_dataInicial__date__lte=data_fim
            )
            .values("RAM_servico__nome")
            .annotate(total_horas=Sum(F("RAM_dataFinal") - F("RAM_dataInicial"))
            )
            .order_by("-total_horas")
        )

        servicos_labels = []
        servicos_values = []
        total_horas = timedelta()
        
        for servico in servicos_atividades:
            duracao_timedelta = servico["total_horas"]
            if duracao_timedelta:
                total_horas += duracao_timedelta
                servicos_labels.append(servico["RAM_servico__nome"])
                servicos_values.append(round(duracao_timedelta.total_seconds() / 3600, 2))  # Converte para horas


        def formatar_horas(duracao):
            if isinstance(duracao, timedelta):
                total_segundos = int(duracao.total_seconds())
            else:
                total_segundos = int(float(duracao) * 3600)

            horas, resto = divmod(total_segundos, 3600)
            minutos, segundos = divmod(resto, 60)

            return f"{horas}:{minutos:02}:{segundos:02}" 

        total_horas_formatado = formatar_horas(total_horas)

        # ---- Gráfico de Barras Empilhadas (Serviços por Dia) ----
        registros_por_dia_servico = (
            RegistroAtividadeModel.objects
            .filter(
                RAM_colaborador__groups__name__in=setores_dinamicos,
                RAM_dataInicial__date__gte=data_inicio,
                RAM_dataInicial__date__lte=data_fim
            )
            .values("RAM_dataInicial__date", "RAM_servico__nome")
            .annotate(total_horas=Sum(F("RAM_dataFinal") - F("RAM_dataInicial")))
            .order_by("RAM_dataInicial__date", "RAM_servico__nome")
        )

        dados_agrupados = {}
        for registro in registros_por_dia_servico:
            dia = registro["RAM_dataInicial__date"].strftime("%d/%m")
            servico = registro["RAM_servico__nome"]
            if registro["total_horas"] is not None:
                horas = registro["total_horas"].total_seconds() / 3600
            else:
                horas = 0

            if horas > 0:
                if dia not in dados_agrupados:
                    dados_agrupados[dia] = {}
                dados_agrupados[dia][servico] = horas

        # Criando as estruturas para o gráfico
        barras_labels = list(dados_agrupados.keys())  # Lista de dias selecionados
        servicos_unicos = sorted(set(servico for dia in dados_agrupados.values() for servico in dia))  # Serviços únicos

        barras_datasets = []
        cores = ["#3498db", "#2ecc71", "#e74c3c", "#f1c40f", "#9b59b6", "#1abc9c", "#ff5733"]  # Lista de cores

        for i, servico in enumerate(servicos_unicos):
            data = [dados_agrupados[dia].get(servico, 0) for dia in barras_labels]  # Horas por dia

            if sum(data) > 0:  # Garante que o serviço tenha ao menos uma entrada maior que zero
                dataset = {
                    "label": servico,
                    "backgroundColor": cores[i % len(cores)],  # Rotaciona as cores
                    "data": data
                }
            barras_datasets.append(dataset)


        # Adicionando ao contexto
        context["total_horas"] = formatar_horas(total_horas)
        context["atividades_por_dia"] = json.dumps(atividades_por_dia)
        context["servicos_labels"] = json.dumps(servicos_labels)
        context["servicos_values"] = json.dumps(servicos_values)
        context["barras_labels"] = json.dumps(barras_labels)
        context["barras_datasets"] = json.dumps(barras_datasets)
        context["data_inicio"] = data_inicio.strftime("%Y-%m-%d")
        context["data_fim"] = data_fim.strftime("%Y-%m-%d")

        return context
    
    
# Função para formatar a duração em "hh:mm:ss"
def formatar_duracao(total_segundos):
    horas = int(total_segundos // 3600)
    minutos = int((total_segundos % 3600) // 60)
    segundos = int(total_segundos % 60)
    return f"{horas}:{minutos:02d}:{segundos:02d}"

class RelatorioHandler:
    @staticmethod
    def criar_pdf(atividades, formatted_total_duracao):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elementos = []
        styles = getSampleStyleSheet()

        data_geracao = datetime.now(pytz.timezone('America/Sao_Paulo')).strftime('%d/%m/%Y %H:%M:%S')
        elementos.append(Paragraph(f"Data de Geração: {data_geracao}", styles['Normal']))
        elementos.append(Spacer(1, 12))
        titulo_style = styles['Title']
        titulo_style.textColor = colors.HexColor("#007bff")
        elementos.append(Paragraph("Relatório de Atividades", titulo_style))
        elementos.append(Spacer(1, 12))

        cell_style = ParagraphStyle(
            name="cell_style",
            fontName="Helvetica",
            fontSize=9,
            leading=12,
            alignment=1,
            wordWrap='LTR'
        )

        dados = [['Colaborador', 'Cliente', 'Serviço', 'Atividade', 'Data Inicial', 'Data Final', 'Duração']]
        for atividade in atividades:
            dados.append([
                str(atividade.RAM_colaborador),
                str(atividade.RAM_cliente),
                str(atividade.RAM_servico),
                Paragraph(str(atividade.RAM_atividade), cell_style),
                atividade.RAM_dataInicial.strftime("%d/%m/%Y %H:%M:%S") if atividade.RAM_dataInicial else '0:00:00',
                atividade.RAM_dataFinal.strftime("%d/%m/%Y %H:%M:%S") if atividade.RAM_dataFinal else '0:00:00',
                atividade.duracao_formatada,
            ])

        tabela = Table(dados, colWidths=[70, 70, 100, 100, 90, 90, 70])
        estilo_tabela = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#007bff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ("ALIGN", (0, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ])
        tabela.setStyle(estilo_tabela)
        elementos.append(tabela)
        elementos.append(Spacer(1, 12))

        total_style = styles['Normal']
        total_style.fontName = "Helvetica-Bold"
        total_style.textColor = colors.HexColor("#343a40")
        elementos.append(Paragraph(f'<b>Total de Duração:</b> {formatted_total_duracao}', total_style))

        doc.build(elementos)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_para_excel(atividades):
        # Cria a lista de dados; para as datas, convertemos os datetimes com fuso para naive.
        data = [
            {
                'Colaborador': str(atividade.RAM_colaborador),
                'Cliente': str(atividade.RAM_cliente),
                'Serviço': str(atividade.RAM_servico),
                'Atividade': str(atividade.RAM_atividade),
                'Data Inicial': make_naive(atividade.RAM_dataInicial) if atividade.RAM_dataInicial else pd.NaT,
                'Data Final': make_naive(atividade.RAM_dataFinal) if atividade.RAM_dataFinal else pd.NaT,
                'Duração': atividade.RAM_duracao if atividade.RAM_duracao else pd.Timedelta(0),
            }
            for atividade in atividades
        ]
        
        # Calcula o total de duração em segundos e converte para timedelta
        total_segundos = sum(
            (atividade.RAM_dataFinal - atividade.RAM_dataInicial).total_seconds()
            for atividade in atividades if atividade.RAM_dataFinal and atividade.RAM_dataInicial
        )
        total_duracao = pd.Timedelta(seconds=total_segundos)
        
        # Adiciona a linha do total de duração
        data.append({
            'Colaborador': '',
            'Cliente': '',
            'Serviço': '',
            'Atividade': '',
            'Data Inicial': None,
            'Data Final': 'Total de Duração:',
            'Duração': total_duracao,
        })
        
        df = pd.DataFrame(data)
        
        # Use ExcelWriter com o engine openpyxl
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Relatorio')
            workbook = writer.book
            worksheet = writer.sheets['Relatorio']
            
            from openpyxl.utils import get_column_letter
            
            # Defina os formatos desejados para as colunas:
            # - Data Inicial e Data Final: formato de data/hora "dd/mm/yyyy hh:mm:ss"
            # - Duração: formato de tempo "[h]:mm:ss"
            formatações = {
                "Data Inicial": "dd/mm/yyyy hh:mm:ss",
                "Data Final": "dd/mm/yyyy hh:mm:ss",
                "Duração": "[h]:mm:ss"
            }
            
            for col_name, num_fmt in formatações.items():
                # Obtenha o índice da coluna (openpyxl é 1-indexado)
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                # Ajusta a largura da coluna (opcional)
                worksheet.column_dimensions[col_letter].width = 20
                # Aplica o formato para todas as células dessa coluna (exceto o cabeçalho)
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{col_letter}{row}"]
                    cell.number_format = num_fmt
                    
        buffer.seek(0)
        return buffer


class GerarRelatorioView(LoginRequiredMixin, AdminRequiredMixin, View):
    login_url = "login/"

    def get(self, request, *args, **kwargs):
        def get_user_groups(user):
            """Retorna os grupos do usuário, excluindo ADMINISTRADOR e USER."""
            return user.groups.exclude(name__in=['ADMINISTRADOR', 'USER'])

        setores_usuario = get_user_groups(request.user)

        atividades = RegistroAtividadeModel.objects.filter(
            RAM_cliente__setor__in=setores_usuario,
            RAM_atividade__setor__in=setores_usuario,
            RAM_colaborador__groups__in=setores_usuario
        ).distinct()

        # Aplicação de Filtros
        servico_ids = request.GET.getlist('servico')
        if servico_ids:
            atividades = atividades.filter(RAM_servico__id__in=servico_ids)

        cliente_ids = request.GET.getlist('cliente')
        if cliente_ids:
            atividades = atividades.filter(RAM_cliente__id__in=cliente_ids)

        atividade_ids = request.GET.getlist('atividade')
        if atividade_ids:
            atividades = atividades.filter(RAM_atividade__id__in=atividade_ids)

        colaborador_ids = request.GET.getlist('colaborador')
        if colaborador_ids:
            atividades = atividades.filter(RAM_colaborador__id__in=colaborador_ids)

        # Filtros de data
        hora = request.GET.get('hora')
        data_fim = request.GET.get('data_fim')
        if hora and data_fim:
            try:
                hora_datetime = datetime.strptime(hora, '%Y-%m-%d')
                data_fim_datetime = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                atividades = atividades.filter(RAM_dataInicial__gte=hora_datetime, RAM_dataInicial__lte=data_fim_datetime)
            except ValueError:
                pass

        # **🚀 Separação das Atividades**
        atividades_abertas = atividades.filter(RAM_dataFinal__isnull=True).order_by('-RAM_duracao')
        atividades_fechadas = atividades.filter(RAM_dataFinal__isnull=False).order_by('-RAM_dataInicial')

        # Combina as listas
        atividades = list(atividades_abertas) + list(atividades_fechadas)

        # Paginação
        paginator = Paginator(atividades, 20)
        atividades_page = paginator.get_page(request.GET.get('page'))

        # Calcula a duração total
        total_segundos = sum(
            (a.RAM_dataFinal - a.RAM_dataInicial).total_seconds()
            for a in atividades if a.RAM_dataFinal and a.RAM_dataInicial
        )
        formatted_total_duracao = formatar_duracao(total_segundos)

        # Exportação de PDF ou Excel
        if 'gerar_pdf' in request.GET:
            buffer = RelatorioHandler.criar_pdf(atividades, formatted_total_duracao)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="relatorio_atividades.pdf"'
            return response

        if 'gerar_excel' in request.GET:
            buffer = RelatorioHandler.exportar_para_excel(atividades)
            response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="relatorio_atividades.xlsx"'
            return response

        # Carregar os dados filtrados
        clientes = Cliente.objects.filter(setor__in=setores_usuario).distinct().order_by('nome')
        servicos = Servico.objects.filter(setor__in=setores_usuario).distinct().order_by('nome')
        atividadegeral = Atividade.objects.filter(setor__in=setores_usuario).distinct().order_by('nome')
        colaboradores = User.objects.filter(groups__in=setores_usuario).distinct().order_by('username')

        context = {
            'atividades': atividades_page,
            'page_obj': atividades_page,
            'clientes': clientes,
            'servicos': servicos,
            'atividadegeral': atividadegeral,
            'colaboradores': colaboradores,
            'hora': hora,
            'data_fim': data_fim,
            'total_duracao': formatted_total_duracao,
            'is_admin': request.user.groups.filter(name='Admin').exists(),
            'colaborador_ids': request.GET.getlist('colaborador'),
            'cliente_ids': request.GET.getlist('cliente'),
            'servico_ids': request.GET.getlist('servico'),
            'atividadegeral_ids': request.GET.getlist('atividade'),
        }
        context.update(aside_icons(request))
        return render(request, 'admin/relatorio.html', context)
